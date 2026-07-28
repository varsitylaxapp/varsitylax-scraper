#!/usr/bin/env python3
"""Parse the WHSBLA league exports into a JSON payload for import-whsbla.js.

Read-only. Touches no database. Run with the scratchpad venv:
    <venv>/bin/python scripts/whsbla-extract.py <out.json>

Shape notes discovered by inspection (they differ from the informal description):
  * Schedule header has DUPLICATE column names:
        ('Date','Time','Away','Home','Away','Home','Location','Type','Notation')
    Columns 2/3 are team NAMES, 4/5 are SCORES. Positional access only —
    name-based lookup silently returns the wrong column.
  * Time cells are Excel time serials rendered as datetime(1900,1,1,H,M).
  * Notation has four values, not two: Overtime, Forfeit, One Referee,
    No Referees.
"""
import openpyxl, json, re, sys, collections, datetime

DIR = 'data/whsbla-2026/'
TEAMS = DIR + 'Teams.xlsx'
SCHED = DIR + '2026-WHSBLA-Varsity-(Lacrosse)-Schedule.xlsx'
TAG = re.compile(r'\s*\(([A-Z]{2})\)\s*$')
DATE, TIME, AWAY, HOME, ASC, HSC, LOC, TYPE, NOTE = range(9)


def norm(s):
    """Normalized key for cross-source name matching."""
    s = s.lower().strip()
    s = re.sub(r'\b(high school|hs|high)\b', '', s)
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def slugify(name, suffix='_wa'):
    """Team slug: snake_case + state suffix (decision D4 — suffix, not prefix)."""
    s = re.sub(r'[^a-z0-9]+', '_', name.lower().strip()).strip('_')
    return s + suffix


def main(out_path):
    payload = {}

    # ── Teams ────────────────────────────────────────────────────────────────
    ws = openpyxl.load_workbook(TEAMS, data_only=True)['Teams']
    raw = [str(r[1]).strip() for r in list(ws.iter_rows(values_only=True))[1:] if r and r[1]]

    seen, dupes, ordered = set(), [], []
    for n in raw:
        if n in seen:
            dupes.append(n)
            continue
        seen.add(n)
        ordered.append(n)

    members = [n for n in ordered if not TAG.search(n)]
    tagged = [{'name': TAG.sub('', n).strip(), 'state': TAG.search(n).group(1), 'raw': n}
              for n in ordered if TAG.search(n)]

    payload['duplicates_dropped'] = dupes
    payload['members'] = [{'name': n, 'slug': slugify(n), 'norm': norm(n)} for n in members]
    payload['tagged'] = tagged

    # ── Schedule ─────────────────────────────────────────────────────────────
    ws2 = openpyxl.load_workbook(SCHED, data_only=True)['WBLA Schedule']
    rows = [r for r in list(ws2.iter_rows(values_only=True))[1:]
            if r and any(v is not None and str(v).strip() for v in r)]

    games = []
    for r in rows:
        d = r[DATE]
        t = r[TIME]
        dt = None
        if isinstance(d, datetime.datetime):
            date_s = d.date().isoformat()
            if isinstance(t, datetime.datetime):
                dt = f'{date_s} {t.hour:02d}:{t.minute:02d}:00'
        else:
            date_s = str(d)
        gtype = (str(r[TYPE]).strip() if r[TYPE] else '') or 'Normal'
        note = str(r[NOTE]).strip() if r[NOTE] else None
        games.append({
            'date': date_s,
            'datetime': dt,
            'away': str(r[AWAY]).strip() if r[AWAY] else None,
            'home': str(r[HOME]).strip() if r[HOME] else None,
            'away_score': r[ASC] if isinstance(r[ASC], int) else None,
            'home_score': r[HSC] if isinstance(r[HSC], int) else None,
            'location': str(r[LOC]).strip() if r[LOC] else None,
            'type': gtype,
            'notation': note,
            'is_overtime': 1 if note and 'overtime' in note.lower() else 0,
            'is_forfeit': 1 if note and 'forfeit' in note.lower() else 0,
            # Exhibition/Practice are imported but excluded from standings math
            # by is_scrimmage, which v_team_season_record already filters on.
            'is_scrimmage': 1 if gtype in ('Exhibition', 'Practice') else 0,
            # 'Normal' = league game; 'NL' = explicitly non-league.
            'is_league': 1 if gtype == 'Normal' else 0,
            'is_playoff': 1 if gtype == 'Playoff' else 0,
        })
    payload['games'] = games

    payload['stats'] = {
        'teams_rows': len(raw),
        'members': len(members),
        'tagged': len(tagged),
        'duplicates': collections.Counter(dupes),
        'games': len(games),
        'types': collections.Counter(g['type'] for g in games),
        'notations': collections.Counter(g['notation'] for g in games),
        'date_min': min(g['date'] for g in games),
        'date_max': max(g['date'] for g in games),
        'missing_scores': sum(1 for g in games
                              if g['away_score'] is None or g['home_score'] is None),
        'tagged_states': collections.Counter(t['state'] for t in tagged),
    }

    with open(out_path, 'w') as f:
        json.dump(payload, f, indent=1, default=str)

    s = payload['stats']
    print(f"  teams rows={s['teams_rows']} members={s['members']} tagged={s['tagged']} "
          f"dupes_dropped={dict(s['duplicates'])}")
    print(f"  games={s['games']} types={dict(s['types'])}")
    print(f"  notations={ {k: v for k, v in s['notations'].items() if k} }")
    print(f"  dates {s['date_min']} -> {s['date_max']}  missing_scores={s['missing_scores']}")
    print(f"  wrote {out_path}")


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else 'whsbla.json')
